"""Bounded, non-persistent profiling for user-provided tabular data."""

from __future__ import annotations

import base64
import binascii
from dataclasses import dataclass
from datetime import datetime, timezone
import csv
import hashlib
from io import BytesIO, StringIO
import json
import math
from pathlib import Path
import re
from typing import Any
import zipfile

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook


MAX_CONTENT_BYTES = 1_000_000
MAX_ENCODED_CONTENT_CHARS = 1_400_000
MAX_XLSX_EXPANDED_BYTES = 20_000_000
MAX_ROWS = 10_000
MAX_COLUMNS = 100
MAX_COLUMN_NAME_LENGTH = 120
MAX_HEADER_SCAN_ROWS = 10
SUPPORTED_FORMATS = ("csv", "json", "xlsx", "parquet")
UNNAMED_COLUMN_PATTERN = re.compile(
    r"^unnamed:\s*\d+(?:_level_\d+)?$",
    re.IGNORECASE,
)
HEADER_CONTROL_CHARACTER_PATTERN = re.compile(r"[\x00-\x1f\x7f]+")
ISO_DATETIME_PATTERN = re.compile(
    r"^\d{4}-\d{2}-\d{2}(?:[T ]\d{2}:\d{2}(?::\d{2}(?:\.\d+)?)?(?:Z|[+-]\d{2}:?\d{2})?)?$"
)
ISO_DATE_ONLY_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ISO_TIMEZONE_SUFFIX_PATTERN = re.compile(r"(?:Z|[+-]\d{2}:?\d{2})$")
SENSITIVE_COLUMN_PATTERN = re.compile(
    r"(^|[\s_-])(password|passwd|secret|api[\s_-]?key|access[\s_-]?token|refresh[\s_-]?token|credential|ssn|social[\s_-]?security|resident[\s_-]?registration|national[\s_-]?id|주민등록번호)([\s_-]|$)",
    re.IGNORECASE,
)
HEADER_NAME_HINTS = (
    "id",
    "name",
    "date",
    "time",
    "year",
    "month",
    "day",
    "count",
    "amount",
    "value",
    "total",
    "category",
    "region",
    "요일",
    "일수",
    "날짜",
    "시간",
    "배출",
    "횟수",
    "수량",
    "금액",
    "합계",
    "평균",
    "구분",
)
SUMMARY_ROW_LABELS = frozenset(
    {"합계", "총계", "소계", "subtotal", "total", "grand total"}
)


class DatasetAnalysisError(ValueError):
    """Raised when an uploaded dataset violates the public-safe contract."""


def summary_row_details(frame: pd.DataFrame) -> tuple[int, list[tuple[str, str]]]:
    """Return summary-row count and unique ``(column, label)`` exclusion pairs."""

    row_count = 0
    column_labels: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    columns = [str(column) for column in frame.columns]
    for row in frame.itertuples(index=False, name=None):
        row_has_summary = False
        for column, value in zip(columns, row, strict=True):
            if value is None:
                continue
            label = str(value).strip()
            normalized = label.casefold()
            if normalized not in SUMMARY_ROW_LABELS:
                continue
            row_has_summary = True
            key = (column, normalized)
            if key not in seen:
                column_labels.append((column, label))
                seen.add(key)
        if row_has_summary:
            row_count += 1
    return row_count, column_labels


@dataclass(frozen=True)
class ColumnNameChange:
    """One deterministic, user-visible repair applied to a source header."""

    position: int
    original: str | None
    normalized: str
    reasons: tuple[str, ...]


@dataclass(frozen=True)
class TableStructureNormalization:
    """Recoverable report-style table structure changes."""

    header_row: int = 1
    preamble_rows_removed: int = 0
    blank_rows_removed: int = 0
    detected_title: str | None = None

    @property
    def applied(self) -> bool:
        return self.header_row != 1 or self.blank_rows_removed > 0


@dataclass(frozen=True)
class LoadedDataset:
    """Request-scoped dataset that is never written to persistent storage."""

    filename: str
    data_format: str
    fingerprint_sha256: str
    frame: pd.DataFrame
    column_name_changes: tuple[ColumnNameChange, ...] = ()
    table_structure: TableStructureNormalization = TableStructureNormalization()


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _safe_filename(filename: str, data_format: str) -> str:
    safe = Path(filename).name.strip() or f"uploaded.{data_format}"
    if len(safe) > 120:
        raise DatasetAnalysisError("filename must be 120 characters or fewer")
    suffix = Path(safe).suffix.lower().lstrip(".")
    if suffix and suffix != data_format:
        raise DatasetAnalysisError("filename extension does not match the declared format")
    return safe


def _parse_json(content: str) -> pd.DataFrame:
    try:
        payload = json.loads(content)
    except (json.JSONDecodeError, RecursionError) as exc:
        raise DatasetAnalysisError("JSON content is invalid") from exc
    if isinstance(payload, dict) and isinstance(payload.get("data"), list):
        payload = payload["data"]
    if isinstance(payload, dict):
        payload = [payload]
    if not isinstance(payload, list) or any(not isinstance(item, dict) for item in payload):
        raise DatasetAnalysisError("JSON must be an object, a list of objects, or an object with a data list")
    if any(
        isinstance(value, (dict, list))
        for item in payload
        for value in item.values()
    ):
        raise DatasetAnalysisError("nested JSON values are not supported; provide flat records")
    return pd.DataFrame(payload)


def _decode_content(content: str, content_encoding: str) -> bytes:
    if not content:
        raise DatasetAnalysisError("dataset content is empty")
    if content_encoding == "utf-8":
        raw = content.encode("utf-8")
    elif content_encoding == "base64":
        try:
            raw = base64.b64decode(content, validate=True)
        except (binascii.Error, ValueError) as exc:
            raise DatasetAnalysisError("base64 dataset content is invalid") from exc
    else:
        raise DatasetAnalysisError("content encoding must be utf-8 or base64")
    if not raw:
        raise DatasetAnalysisError("dataset content is empty")
    if len(raw) > MAX_CONTENT_BYTES:
        raise DatasetAnalysisError(f"dataset content exceeds {MAX_CONTENT_BYTES} bytes")
    return raw


def _is_missing_header(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    try:
        return bool(missing)
    except (TypeError, ValueError):
        return False


def _is_empty_cell(value: Any) -> bool:
    if _is_missing_header(value):
        return True
    return isinstance(value, str) and not value.strip()


def _header_row_stats(values: list[Any]) -> tuple[int, int, int, float]:
    present = [value for value in values if not _is_empty_cell(value)]
    if not present:
        return 0, 0, 0, 0.0
    text_values = [value.strip() for value in present if isinstance(value, str)]
    hints = sum(
        any(hint in value.casefold() for hint in HEADER_NAME_HINTS)
        for value in text_values
    )
    return len(present), len(text_values), hints, len(text_values) / len(present)


def _raw_table_with_source_header(frame: pd.DataFrame) -> pd.DataFrame:
    columns = list(range(len(frame.columns)))
    body = frame.copy()
    body.columns = columns
    source_header = pd.DataFrame([list(frame.columns)], columns=columns)
    return pd.concat([source_header, body], ignore_index=True)


def _detect_header_row(raw_table: pd.DataFrame) -> int:
    """Return a zero-based header row only when a later row is clearly stronger."""

    width = len(raw_table.columns)
    if width < 2 or len(raw_table) < 3:
        return 0
    first_count, first_text, first_hints, _ = _header_row_stats(
        raw_table.iloc[0].tolist()
    )
    first_score = first_count * 2 + first_text + first_hints * 2
    best_index = 0
    best_score = first_score
    scan_limit = min(len(raw_table) - 1, MAX_HEADER_SCAN_ROWS)

    for index in range(1, scan_limit):
        values = raw_table.iloc[index].tolist()
        count, text_count, hints, text_ratio = _header_row_stats(values)
        if count < max(2, math.ceil(width * 0.6)) or text_ratio < 0.6:
            continue
        if count < first_count + 2:
            continue
        blank_before = any(
            _header_row_stats(raw_table.iloc[prior].tolist())[0] == 0
            for prior in range(index)
        )
        if not (first_count == 1 or blank_before or hints >= 2):
            continue
        if not any(
            _header_row_stats(raw_table.iloc[after].tolist())[0] > 0
            for after in range(index + 1, len(raw_table))
        ):
            continue
        score = count * 2 + text_count + hints * 2
        if score >= first_score + 6 and score > best_score:
            best_index = index
            best_score = score
    return best_index


def _detected_table_title(raw_table: pd.DataFrame, header_index: int) -> str | None:
    for index in range(header_index):
        present = [value for value in raw_table.iloc[index].tolist() if not _is_empty_cell(value)]
        if len(present) != 1 or not isinstance(present[0], str):
            continue
        title = " ".join(present[0].split())
        if title:
            return title[:MAX_COLUMN_NAME_LENGTH]
    return None


def _normalize_table_structure(
    frame: pd.DataFrame,
    data_format: str,
) -> tuple[pd.DataFrame, TableStructureNormalization]:
    if data_format not in {"csv", "xlsx"}:
        return frame, TableStructureNormalization()

    raw_table = _raw_table_with_source_header(frame)
    header_index = _detect_header_row(raw_table)
    if header_index:
        prepared = raw_table.iloc[header_index + 1 :].copy()
        prepared.columns = raw_table.iloc[header_index].tolist()
    else:
        prepared = frame.copy()

    blank_rows = prepared.apply(
        lambda row: all(_is_empty_cell(value) for value in row.tolist()),
        axis=1,
    )
    blank_rows_removed = int(blank_rows.sum())
    prepared = prepared.loc[~blank_rows].reset_index(drop=True)
    return prepared, TableStructureNormalization(
        header_row=header_index + 1,
        preamble_rows_removed=header_index,
        blank_rows_removed=blank_rows_removed,
        detected_title=_detected_table_title(raw_table, header_index),
    )


def _infer_scalar_types(frame: pd.DataFrame) -> pd.DataFrame:
    inferred = frame.copy()
    for column in inferred.columns:
        series = inferred[column].map(
            lambda value: pd.NA if isinstance(value, str) and not value.strip() else value
        )
        present_count = int(series.notna().sum())
        if not present_count:
            inferred[column] = series
            continue
        numeric = pd.to_numeric(series, errors="coerce")
        inferred[column] = numeric if int(numeric.notna().sum()) == present_count else series
    return inferred


def _normalize_column_names(
    header: list[Any],
) -> tuple[list[str], tuple[ColumnNameChange, ...]]:
    """Repair recoverable header defects without changing column order or values."""

    normalized_names: list[str] = []
    changes: list[ColumnNameChange] = []
    seen: dict[str, str] = {}

    for position, value in enumerate(header, start=1):
        missing = _is_missing_header(value)
        original = None if missing else str(value)
        reasons: list[str] = []
        if missing:
            text = ""
        elif isinstance(value, float) and math.isfinite(value) and value.is_integer():
            text = str(int(value))
            reasons.append("numeric_normalized")
        else:
            text = str(value)

        cleaned = HEADER_CONTROL_CHARACTER_PATTERN.sub(" ", text)
        cleaned = " ".join(cleaned.split())
        if cleaned != text:
            reasons.append("whitespace_normalized")
        if not isinstance(value, str) and not missing:
            reasons.append("converted_to_text")

        if not cleaned:
            base = f"column_{position}"
            reasons.append("empty")
        elif UNNAMED_COLUMN_PATTERN.fullmatch(cleaned):
            base = f"column_{position}"
            reasons.append("placeholder_replaced")
        else:
            base = cleaned

        candidate = base[:MAX_COLUMN_NAME_LENGTH]
        if candidate != base:
            reasons.append("truncated")

        counter = 2
        collision_reason: str | None = None
        while candidate.casefold() in seen:
            if collision_reason is None:
                collision_reason = (
                    "duplicate"
                    if seen[candidate.casefold()] == candidate
                    else "case_collision"
                )
            suffix = f"_{counter}"
            candidate = base[: MAX_COLUMN_NAME_LENGTH - len(suffix)] + suffix
            counter += 1
        if collision_reason is not None:
            reasons.append(collision_reason)

        seen[candidate.casefold()] = candidate
        normalized_names.append(candidate)
        if candidate != text or reasons:
            changes.append(
                ColumnNameChange(
                    position=position,
                    original=original,
                    normalized=candidate,
                    reasons=tuple(dict.fromkeys(reasons)),
                )
            )

    return normalized_names, tuple(changes)


def _validate_source_header_safety(header: list[Any]) -> None:
    sensitive: list[str] = []
    for value in header:
        if _is_missing_header(value):
            continue
        cleaned = HEADER_CONTROL_CHARACTER_PATTERN.sub(" ", str(value))
        cleaned = " ".join(cleaned.split())
        if SENSITIVE_COLUMN_PATTERN.search(cleaned):
            sensitive.append(cleaned[:MAX_COLUMN_NAME_LENGTH])
    if sensitive:
        raise DatasetAnalysisError(
            "potential credential columns are not accepted: " + ", ".join(sensitive[:5])
        )


def _restore_source_header(frame: pd.DataFrame, header: list[Any], data_format: str) -> pd.DataFrame:
    if not header or len(frame.columns) != len(header):
        raise DatasetAnalysisError(f"{data_format.upper()} header could not be parsed")
    frame.columns = header
    return frame


def _parse_csv(raw: bytes) -> pd.DataFrame:
    try:
        content = raw.decode("utf-8-sig")
        rows = list(csv.reader(StringIO(content)))
        if not rows:
            raise DatasetAnalysisError("CSV content could not be parsed")
        width = max(len(row) for row in rows)
        if width > MAX_COLUMNS:
            raise DatasetAnalysisError(f"dataset exceeds the {MAX_COLUMNS}-column limit")
        padded = [row + [None] * (width - len(row)) for row in rows]
        return pd.DataFrame(padded[1:], columns=padded[0])
    except DatasetAnalysisError:
        raise
    except (csv.Error, StopIteration, pd.errors.ParserError, UnicodeError, ValueError) as exc:
        raise DatasetAnalysisError("CSV content could not be parsed") from exc


def _parse_xlsx(raw: bytes) -> pd.DataFrame:
    stream = BytesIO(raw)
    try:
        if not zipfile.is_zipfile(stream):
            raise DatasetAnalysisError("XLSX content could not be parsed")
        stream.seek(0)
        with zipfile.ZipFile(stream) as archive:
            if sum(item.file_size for item in archive.infolist()) > MAX_XLSX_EXPANDED_BYTES:
                raise DatasetAnalysisError("XLSX expanded content exceeds the safety limit")
        stream.seek(0)
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            header = list(next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True)))
        finally:
            workbook.close()
        stream.seek(0)
        frame = pd.read_excel(
            stream,
            engine="openpyxl",
            nrows=MAX_ROWS + MAX_HEADER_SCAN_ROWS + 1,
        )
        return _restore_source_header(frame, header, "xlsx")
    except DatasetAnalysisError:
        raise
    except (OSError, StopIteration, ValueError, zipfile.BadZipFile) as exc:
        raise DatasetAnalysisError("XLSX content could not be parsed") from exc


def _parse_parquet(raw: bytes) -> pd.DataFrame:
    stream = BytesIO(raw)
    try:
        parquet = pq.ParquetFile(stream)
        if parquet.metadata.num_rows > MAX_ROWS:
            raise DatasetAnalysisError(f"dataset exceeds the {MAX_ROWS}-row limit")
        schema = parquet.schema_arrow
        if len(schema.names) > MAX_COLUMNS:
            raise DatasetAnalysisError(f"dataset exceeds the {MAX_COLUMNS}-column limit")
        if any(pa.types.is_nested(field.type) for field in schema):
            raise DatasetAnalysisError("nested Parquet columns are not supported")
        return parquet.read().to_pandas()
    except DatasetAnalysisError:
        raise
    except (OSError, ValueError, pa.ArrowException) as exc:
        raise DatasetAnalysisError("Parquet content could not be parsed") from exc


def _parse_content(raw: bytes, data_format: str, content_encoding: str) -> pd.DataFrame:
    if data_format in {"xlsx", "parquet"} and content_encoding != "base64":
        raise DatasetAnalysisError(f"{data_format.upper()} content requires base64 encoding")
    if data_format == "csv":
        return _parse_csv(raw)
    if data_format == "json":
        try:
            return _parse_json(raw.decode("utf-8-sig"))
        except UnicodeDecodeError as exc:
            raise DatasetAnalysisError("JSON content must be UTF-8") from exc
    if data_format == "xlsx":
        return _parse_xlsx(raw)
    if data_format == "parquet":
        return _parse_parquet(raw)
    raise DatasetAnalysisError("dataset format must be csv, json, xlsx, or parquet")


def _validate_shape(frame: pd.DataFrame) -> None:
    if frame.empty:
        raise DatasetAnalysisError("dataset must include at least one data row")
    if len(frame) > MAX_ROWS:
        raise DatasetAnalysisError(f"dataset exceeds the {MAX_ROWS}-row limit")
    if len(frame.columns) > MAX_COLUMNS:
        raise DatasetAnalysisError(f"dataset exceeds the {MAX_COLUMNS}-column limit")
    if len(frame.columns) == 0:
        raise DatasetAnalysisError("dataset must include at least one column")
    if len({str(column).casefold() for column in frame.columns}) != len(frame.columns):
        raise DatasetAnalysisError("normalized dataset column names must be unique")
    sensitive = [str(column) for column in frame.columns if SENSITIVE_COLUMN_PATTERN.search(str(column))]
    if sensitive:
        raise DatasetAnalysisError(
            "potential credential columns are not accepted: " + ", ".join(sensitive[:5])
        )


def _coerce_iso_datetimes(frame: pd.DataFrame) -> pd.DataFrame:
    converted = frame.copy()
    for column in converted.columns:
        series = converted[column]
        if not (pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series)):
            continue
        present = series.dropna().astype(str).str.strip()
        if present.empty or not present.map(lambda value: bool(ISO_DATETIME_PATTERN.fullmatch(value))).all():
            continue
        if present.map(lambda value: bool(ISO_DATE_ONLY_PATTERN.fullmatch(value))).all():
            parsed = pd.to_datetime(series, format="%Y-%m-%d", errors="coerce")
        else:
            timezone_flags = present.map(
                lambda value: bool(ISO_TIMEZONE_SUFFIX_PATTERN.search(value))
            )
            if timezone_flags.nunique() > 1:
                continue
            parsed = pd.to_datetime(
                series,
                format="ISO8601",
                errors="coerce",
                utc=bool(timezone_flags.all()),
            )
        if parsed.notna().sum() == series.notna().sum():
            converted[column] = parsed
    return converted


def _finite_number(value: Any) -> float | int | None:
    if value is None or pd.isna(value):
        return None
    number = float(value)
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else round(number, 6)


def _column_profile(series: pd.Series) -> dict[str, Any]:
    missing_count = int(series.isna().sum())
    profile: dict[str, Any] = {
        "name": str(series.name),
        "dtype": str(series.dtype),
        "missing_count": missing_count,
        "missing_rate": round(missing_count / len(series), 6),
        "unique_count": int(series.nunique(dropna=True)),
    }
    if pd.api.types.is_numeric_dtype(series):
        numeric = pd.to_numeric(series, errors="coerce")
        profile["numeric"] = {
            "count": int(numeric.count()),
            "min": _finite_number(numeric.min()),
            "q1": _finite_number(numeric.quantile(0.25)),
            "max": _finite_number(numeric.max()),
            "mean": _finite_number(numeric.mean()),
            "median": _finite_number(numeric.median()),
            "q3": _finite_number(numeric.quantile(0.75)),
            "stddev": _finite_number(numeric.std()),
        }
    if pd.api.types.is_datetime64_any_dtype(series):
        present = series.dropna()
        profile["temporal"] = {
            "min": present.min().isoformat() if not present.empty else None,
            "max": present.max().isoformat() if not present.empty else None,
        }
    return profile


def load_dataset(
    filename: str,
    data_format: str,
    content: str,
    content_encoding: str = "utf-8",
) -> LoadedDataset:
    """Parse and validate one request-scoped dataset."""
    normalized_format = data_format.strip().lower()
    if normalized_format not in SUPPORTED_FORMATS:
        raise DatasetAnalysisError("dataset format must be csv, json, xlsx, or parquet")
    safe_filename = _safe_filename(filename, normalized_format)
    raw = _decode_content(content, content_encoding.strip().lower())
    frame = _parse_content(raw, normalized_format, content_encoding.strip().lower())
    _validate_source_header_safety(list(frame.columns))
    frame, table_structure = _normalize_table_structure(frame, normalized_format)
    _validate_source_header_safety(list(frame.columns))
    normalized_columns, column_name_changes = _normalize_column_names(list(frame.columns))
    frame.columns = normalized_columns
    if normalized_format in {"csv", "xlsx"}:
        frame = _infer_scalar_types(frame)
    frame = _coerce_iso_datetimes(frame)
    _validate_shape(frame)
    return LoadedDataset(
        filename=safe_filename,
        data_format=normalized_format,
        fingerprint_sha256=hashlib.sha256(raw).hexdigest(),
        frame=frame,
        column_name_changes=column_name_changes,
        table_structure=table_structure,
    )


def profile_dataset(dataset: LoadedDataset) -> dict[str, Any]:
    """Return a public-safe profile without retaining raw user content."""

    frame = dataset.frame
    columns = [_column_profile(frame[column]) for column in frame.columns]
    missing_cells = sum(item["missing_count"] for item in columns)
    numeric_columns = sum("numeric" in item for item in columns)
    duplicate_rows = int(frame.duplicated().sum())
    return {
        "contract_version": "decisionops-dataset-profile-v1",
        "filename": dataset.filename,
        "format": dataset.data_format,
        "fingerprint_sha256": dataset.fingerprint_sha256,
        "generated_at": _utc_now(),
        "row_count": int(len(frame)),
        "column_count": int(len(frame.columns)),
        "numeric_column_count": numeric_columns,
        "missing_cell_count": missing_cells,
        "missing_cell_rate": round(missing_cells / frame.size, 6),
        "duplicate_row_count": duplicate_rows,
        "duplicate_row_rate": round(duplicate_rows / len(frame), 6),
        "columns": columns,
        "column_name_normalization": {
            "applied": bool(dataset.column_name_changes),
            "change_count": len(dataset.column_name_changes),
            "changes": [
                {
                    "position": change.position,
                    "original": change.original,
                    "normalized": change.normalized,
                    "reasons": list(change.reasons),
                }
                for change in dataset.column_name_changes
            ],
        },
        "table_structure_normalization": {
            "applied": dataset.table_structure.applied,
            "header_row": dataset.table_structure.header_row,
            "preamble_rows_removed": dataset.table_structure.preamble_rows_removed,
            "blank_rows_removed": dataset.table_structure.blank_rows_removed,
            "detected_title": dataset.table_structure.detected_title,
        },
        "storage": "not_persisted",
        "limits": {
            "max_content_bytes": MAX_CONTENT_BYTES,
            "max_rows": MAX_ROWS,
            "max_columns": MAX_COLUMNS,
            "supported_formats": list(SUPPORTED_FORMATS),
        },
    }


def analyze_dataset(
    filename: str,
    data_format: str,
    content: str,
    content_encoding: str = "utf-8",
) -> dict[str, Any]:
    """Parse and profile a dataset without retaining raw user content."""

    return profile_dataset(load_dataset(filename, data_format, content, content_encoding))
